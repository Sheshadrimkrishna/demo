import openpyxl
import pytest

def get_data():

    final_list = []
    workbook=openpyxl.load_workbook("ExcelFile/TutorialsNinja.xlsx")
    sheet = workbook['LoginTest']
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    # for r in range(2,total_rows+1):
    #     row_list = []
    #     for c in range(1,total_cols+1):
    #         row_list.append(sheet.cell(row=r,column=c).value)
    #     final_list.append(row_list)
    #
    # return final_list
    for r in range(2, total_rows + 1):
        username = sheet.cell(row=r, column=1).value
        password = sheet.cell(row=r, column=2).value
        if username and password:  # Ensure both username and password are not None or empty
            final_list.append((username, password))

    return final_list

@pytest.mark.parametrize("username,password",get_data())
def test_login(username,password):
    print("Logged in using username: "+username+" and password" +password)



